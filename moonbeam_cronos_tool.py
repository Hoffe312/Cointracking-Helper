import pandas as pd
import openpyxl as op
import csv


#to do:
#if deposit asset == withdraw asset although both hashes are the same its a deposit and withdraw and not a trade

def read_muster():
    df = pd.read_excel('./CoinTracking_Excel_Import.xls')
    row_arr = df.iloc[0].values
    row_arr.tolist()
    type_arr = df.iloc[:, 0].tolist()
    type_arr = type_arr[1:]

    value_dict = {}
    for i in range(len(type_arr)):
        value_dict[type_arr[i]]= df.iloc[i+1,:].tolist()
    return row_arr, type_arr, value_dict

'''Funktion, welche für Abfragen, der unterschiedlichen Börsen verantwortlich ist'''
def path_format_handler():
    #path = input("Gebe den Pfad zu deiner Datei an, welche du in das Cointracking_Excel_Import Format bringen willst:\n")
    path= "C:\\Users\\felix\\OneDrive\\Desktop\\Sieber Python\\koinly_2022 Dr.Sieber.csv"
    data_format=""
    pattern_csv = ".csv"
    pattern_xsl = ".xsl"
    if pattern_csv in path:
        data_format = "c"

    if pattern_xsl in path:
        data_format = "x"
    return path, data_format


'''Liest excel oder csv Dateien ein und gibt dataframes zurück'''
def read_data(path, format):
    if path[0] == '"' and path[-1] == '"':
        path = path[1:-1]
    if format == "x":
        df = pd.read_excel(path)
    elif format == "c":
        with open(path, 'r') as file:
            reader = csv.reader(file, delimiter=',')
            next(reader)  # Überspringe die erste Zeile (Header)
            next(reader)  # Überspringe die zweite Zeile
            data = [row for row in reader]
    print(data[1])
    return data


def save_to_xsl(general,row):
    wb = op.Workbook()
    ws = wb.active
  
    for i, wert in enumerate(row):
        ws.cell(row=1, column=i+1).value = wert
    for i, row in enumerate(general):
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1).value = value
         
    wb.save("C:\\Users\\felix\\source\\repos\\Cointracking_Helper\\Example.xlsx")
    print("******************************************")
    print("*DIE ERSTELLUNG DER DATEI WAR ERFOLGREICH*")
    print("******************************************")

def calculate_num(num):
    if num == '':
        return 0
    zahl_string = num
    zahl_float = float(zahl_string.replace(",","."))
    zahl_abs = abs(zahl_float)
    return zahl_abs


'''Formatiert bitcoin.de dateien'''
def format_data_bitcoinde(row):
    path, format = path_format_handler()
    data = read_data(path,format)

    length_list = len(data)
    general = []
    #if general_array:
    #    general = general_array
    for i in range(length_list):
        if data[i][1] == "crypto_deposit" and data[i][16] == "0x54e2d14df9348b3fba7e372328595b9f3ae243fe" and data[i+1][17] == "0x54e2d14df9348b3fba7e372328595b9f3ae243fe":
            stella_depo_arr = []
            stella_withdraw_arr = []
            wallet_depo_arr = []
            wallet_withdraw_arr = []
            for j in range(len(row)):
                if j == 0:
                    stella_depo_arr.append("Deposit")
                    stella_withdraw_arr.append("Withdrawal")
                    wallet_depo_arr.append("Deposit")
                    wallet_withdraw_arr.append("Withdrawal")
                if j == 1:
                    stella_withdraw_arr.append("")
                    stella_depo_arr.append(calculate_num(data[i][8]))
                    wallet_depo_arr.append(calculate_num(data[i+1][4]))
                    wallet_withdraw_arr.append("")
                if j == 2:
                    stella_withdraw_arr.append("")
                    stella_depo_arr.append(data[i][9])
                    wallet_depo_arr.append(data[i+1][5])
                    wallet_withdraw_arr.append("")
                if j == 3:
                    stella_withdraw_arr.append(calculate_num(data[i+1][4]))
                    stella_depo_arr.append("")
                    wallet_depo_arr.append("")
                    wallet_withdraw_arr.append(calculate_num(data[i][8]))
                if j == 4:
                    stella_withdraw_arr.append(data[i+1][5])
                    stella_depo_arr.append("")
                    wallet_depo_arr.append("")
                    wallet_withdraw_arr.append(data[i][9])
                if j == 5:
                    stella_withdraw_arr.append(calculate_num(data[i+2][4]))
                    stella_depo_arr.append("")
                    wallet_depo_arr.append("")
                    wallet_withdraw_arr.append("")
                if j == 6:
                    stella_withdraw_arr.append(data[i+2][5])
                    stella_depo_arr.append("")
                    wallet_depo_arr.append("")
                    wallet_withdraw_arr.append("")
                if j == 7:
                    stella_withdraw_arr.append(data[i+1][3])
                    stella_depo_arr.append(data[i][7])
                    wallet_depo_arr.append("Stella Vault")
                    wallet_withdraw_arr.append("Stella Vault")
                if j == 8 or j == 9:
                    stella_withdraw_arr.append("")
                    stella_depo_arr.append("")
                    wallet_depo_arr.append("")
                    wallet_withdraw_arr.append("")
                if j == 10:
                    stella_withdraw_arr.append(data[i][0])
                    stella_depo_arr.append(data[i][0])
                    wallet_depo_arr.append(data[i][0])
                    wallet_withdraw_arr.append(data[i][0])
            general.append(stella_depo_arr)
            general.append(stella_withdraw_arr)
            general.append(wallet_depo_arr)
            general.append(wallet_withdraw_arr)
        elif i < length_list - 2 and data[i][1] == "crypto_deposit" and data[i+1][1] == "crypto_withdrawal"  and data[i][18] == data[i+1][18] and data[i+1][2] != "Cost":
            sell_arr = []
            for j in range(len(row)):
                if j == 0:
                    sell_arr.append("Trade")
                elif j == 1:
                    sell_arr.append(calculate_num(data[i][8]))
                elif j == 2:
                    sell_arr.append(data[i][9])
                elif j == 3:
                    sell_arr.append(calculate_num(data[i+1][4]))
                elif j == 4:
                    sell_arr.append(data[i+1][5])
                elif j == 5:
                    if data[i+2][18] == data[i][18] and data[i+2][2] == "Cost":
                        sell_arr.append(calculate_num(data[i+2][4]))
                    else:
                        sell_arr.append("")
                elif j == 6:
                    if data[i+2][18] == data[i][18] and data[i+2][2] == "Cost":
                        sell_arr.append(data[i+2][5])
                    else:
                        sell_arr.append("")
                elif j == 7:
                    sell_arr.append(data[i][7])
                elif j == 8 or j == 9:
                    sell_arr.append("")
                elif j == 10:
                    sell_arr.append(data[i][0])
            general.append(sell_arr)
        
        elif str(data[i][1]) == "crypto_withdrawal" and data[i-1][18]!=data[i][18] and data[i][18]==data[i+1][18] and data[i+1][2]=="Cost" and i<length_list-1:
            withdraw_arr = [] 
            for j in range(len(row)):
                if j == 0:
                    withdraw_arr.append("Withdrawal")
                if j == 1 or j == 2 or  j == 8 or j == 9:
                    withdraw_arr.append("")
                if j == 5 :                       
                    withdraw_arr.append(calculate_num(data[i+1][4]))
                if j == 6:
                    withdraw_arr.append(data[i+1][5])
                if row[j] == "Sell Amount":
                    withdraw_arr.append(calculate_num(data[i][4]))
                if row[j] == "Sell Cur.":
                    withdraw_arr.append(data[i][5])
                if row[j] == "Exchange (optional)":
                    withdraw_arr.append(data[i][3])
                if row[j] == "Date":
                    withdraw_arr.append(data[i][0])
            general.append(withdraw_arr)
        elif str(data[i][1]) == "crypto_withdrawal" and data[i-1][18]==data[i][18]  and data[i][2]=="Cost" and data[i-1][1] == "crypto_deposit" :
            withdraw_arr = [] 
            for j in range(len(row)):
                if j == 0:
                    withdraw_arr.append("Other Fee")
                if j == 1 or j == 2 or  j == 8 or j == 9:
                    withdraw_arr.append("")
                if j == 5 :                       
                    withdraw_arr.append("")
                if j == 6:
                    withdraw_arr.append("")
                if row[j] == "Sell Amount":
                    withdraw_arr.append(calculate_num(data[i][4]))
                if row[j] == "Sell Cur.":
                    withdraw_arr.append(data[i][5])
                if row[j] == "Exchange (optional)":
                    withdraw_arr.append(data[i][3])
                if row[j] == "Date":
                    withdraw_arr.append(data[i][0])
            general.append(withdraw_arr)
        elif str(data[i][1]) == "crypto_deposit"  and data[i][2] != "Reward" and data[i][18]!=data[i+1][18]:
            depo_arr = [] 
            for j in range(len(row)):
                if j == 0:
                    depo_arr.append("Deposit")
                if j == 3 or j == 4 or j ==5 or j == 6 or  j == 8 or j == 9:
                    depo_arr.append("")
                if row[j] == "Buy Amount":
                    depo_arr.append(calculate_num(data[i][8]))
                if row[j] == "Buy Cur.":
                    depo_arr.append(data[i][9])
                if row[j] == "Exchange (optional)":
                    depo_arr.append(data[i][7])
                if row[j] == "Date":
                    depo_arr.append(data[i][0])
            general.append(depo_arr)
        elif str(data[i][1]) == "crypto_deposit" and data[i][18]!=data[i+1][18] and data[i][2] == "Reward":
            depo_arr = [] 
            for j in range(len(row)):
                if j == 0:
                    depo_arr.append("Reward / Bonus")
                if j == 3 or j == 4 or j ==5 or j == 6 or  j == 8 or j == 9:
                    depo_arr.append("")
                if row[j] == "Buy Amount":
                    zahl_string = data[i][4]
                    zahl_float = float(zahl_string.replace(",","."))
                    zahl_abs = abs(zahl_float)
                    depo_arr.append(zahl_abs)
                if row[j] == "Buy Cur.":
                    depo_arr.append(data[i][5])
                if row[j] == "Exchange (optional)":
                    depo_arr.append(data[i][3])
                if row[j] == "Date":
                    depo_arr.append(data[i][0])
            general.append(depo_arr)
        elif str(data[i][1]) == "crypto_deposit" and data[i][18]==data[i+1][18]==data[i+2][18]==data[i+3][18] and data[i+1][1]=="crypto_withdrawal" and data[i+2][1]=="crypto_withdrawal":
            depo_arr2 = [] #income non taxable
            withdraw_arr1 = []
            withdraw_arr2 = []
            deposit_arr1=[]
            deposit_arr2=[]
            for j in range(len(row)):
                if j == 0:
                    depo_arr2.append("Income (non taxable)")
                    withdraw_arr1.append("Withdrawal")
                    withdraw_arr2.append("Withdrawal")
                    deposit_arr1.append("Deposit")
                    deposit_arr2.append("Deposit")
                if j == 8 or j == 9:
                    depo_arr2.append("")
                    withdraw_arr1.append("")
                    withdraw_arr2.append("")
                    deposit_arr1.append("")
                    deposit_arr2.append("")
                if j == 5 :                       
                    withdraw_arr1.append(calculate_num(data[i+1][4]))
                    withdraw_arr2.append("")
                    deposit_arr1.append("")
                    deposit_arr2.append("")
                    depo_arr2.append("")
                if j == 6:
                    withdraw_arr1.append(data[i+3][5])
                    withdraw_arr2.append("")
                    deposit_arr1.append("")
                    deposit_arr2.append("")
                    depo_arr2.append("")
                if row[j] == "Sell Amount":#3
                    withdraw_arr1.append(calculate_num(data[i+1][4]))
                    withdraw_arr2.append(calculate_num(data[i+2][4]))
                    deposit_arr1.append("")
                    deposit_arr2.append("")
                    depo_arr2.append("")
                if row[j] == "Sell Cur.":#4
                    withdraw_arr1.append(data[i+2][5])
                    withdraw_arr2.append(data[i+3][5])
                    deposit_arr1.append("")
                    deposit_arr2.append("")
                    depo_arr2.append("")
                if row[j] == "Buy Amount":#1
                    depo_arr2.append(calculate_num(data[i][8]))
                    withdraw_arr1.append("")
                    withdraw_arr2.append("")
                    deposit_arr1.append(calculate_num(data[i+1][4]))
                    deposit_arr2.append(calculate_num(data[i+2][4]))
                if row[j] == "Buy Cur.":#2
                    depo_arr2.append(data[i][9])
                    withdraw_arr1.append("")
                    withdraw_arr2.append("")
                    deposit_arr1.append(data[i+1][5])
                    deposit_arr2.append(data[i+2][5])
                if row[j] == "Exchange (optional)":
                    depo_arr2.append("Uniswap V2 "+data[i][9])
                    withdraw_arr1.append(data[i+1][3])
                    withdraw_arr2.append(data[i+1][3])
                    deposit_arr1.append("Uniswap V2 "+data[i][9])
                    deposit_arr2.append("Uniswap V2 "+data[i][9])
                if row[j] == "Date":
                    depo_arr2.append(data[i][0])
                    withdraw_arr1.append(data[i][0])
                    withdraw_arr2.append(data[i][0])
                    deposit_arr1.append(data[i][0])
                    deposit_arr2.append(data[i][0])
            general.append(depo_arr2)
            general.append(withdraw_arr1)
            general.append(withdraw_arr2)
            general.append(deposit_arr1)
            general.append(deposit_arr2)
        elif data[i][1] == "exchange" and data[i][2] == "":
            trade_arr = []
            for j in range(len(row)):
                if j == 0:
                    trade_arr.append("Trade")
                if j == 1:
                    trade_arr.append(calculate_num(data[i][8]))
                if j == 2:
                   trade_arr.append(data[i][9])
                if j == 3:
                    trade_arr.append(calculate_num(data[i][4]))
                if j == 4:
                    trade_arr.append(data[i][5])
                if j == 5:
                    trade_arr.append(calculate_num(data[i][11]))
                if j == 6:
                    trade_arr.append(data[i][12])
                if j == 7:
                    trade_arr.append(data[i][3])
                if j == 8 or j == 9:
                    trade_arr.append("")
                if j == 10:
                    trade_arr.append(data[i][0])
            general.append(trade_arr)
        elif data[i][1] == "exchange" and data[i][2] == "Liquidity in":
            income_non_tax_arr = []
            withdraw2_arr = []
            deposit2_arr = []
            for j in range(len(row)):
                if j == 0:
                    deposit2_arr.append("Deposit")
                    withdraw2_arr.append("Withdrawal")
                    income_non_tax_arr.append("Income (non taxable)")
                if  j == 8 or j == 9:
                    deposit2_arr.append("")
                    withdraw2_arr.append("")
                    income_non_tax_arr.append("")
                if row[j] == "Buy Amount":
                    deposit2_arr.append(calculate_num(data[i][4]))
                    withdraw2_arr.append("")
                    income_non_tax_arr.append(calculate_num(data[i][8]))
                if row[j] == "Buy Cur.":
                    deposit2_arr.append(data[i][5])
                    withdraw2_arr.append("")
                    income_non_tax_arr.append(data[i][9])
                if row[j] == "Sell Amount":
                    deposit2_arr.append("")
                    withdraw2_arr.append(calculate_num(data[i][4]))
                    income_non_tax_arr.append("")
                if j == 4:
                    deposit2_arr.append("")
                    withdraw2_arr.append(data[i][5])
                    income_non_tax_arr.append("")
                if j == 5:
                    deposit2_arr.append("")
                    withdraw2_arr.append(calculate_num(data[i][12]))
                    income_non_tax_arr.append("")
                if j == 6:
                    deposit2_arr.append("")
                    withdraw2_arr.append(data[i][13])
                    income_non_tax_arr.append("")
                if row[j] == "Exchange (optional)":
                    deposit2_arr.append("Uniswap V2 "+data[i][9])
                    withdraw2_arr.append(data[i][3])
                    income_non_tax_arr.append("Uniswap V2 "+data[i][9])
                if row[j] == "Date":
                    deposit2_arr.append(data[i][0])
                    withdraw2_arr.append(data[i][0])
                    income_non_tax_arr.append(data[i][0])
            general.append(withdraw2_arr)
            general.append(deposit2_arr)
            general.append(income_non_tax_arr)
        elif data[i][1] == "exchange" and data[i][2] == "Liquidity out":
            expense_non_tax_arr = []
            withdraw2_arr = []
            deposit2_arr = []
            for j in range(len(row)):
                if j == 0:
                    deposit2_arr.append("Deposit")
                    withdraw2_arr.append("Withdrawal")
                    expense_non_tax_arr.append("Expense (non taxable)")
                if  j == 8 or j == 9:
                    deposit2_arr.append("")
                    withdraw2_arr.append("")
                    expense_non_tax_arr.append("")
                if row[j] == "Buy Amount":
                    deposit2_arr.append(calculate_num(data[i][8]))
                    withdraw2_arr.append("")
                    expense_non_tax_arr.append(calculate_num(data[i][4]))
                if row[j] == "Buy Cur.":
                    deposit2_arr.append(data[i][9])
                    withdraw2_arr.append("")
                    expense_non_tax_arr.append(data[i][5])
                if row[j] == "Sell Amount":
                    deposit2_arr.append("")
                    withdraw2_arr.append(calculate_num(data[i][8]))
                    expense_non_tax_arr.append(calculate_num(data[i][4]))
                if j == 5:
                    deposit2_arr.append("")
                    withdraw2_arr.append("")
                    expense_non_tax_arr.append("")
                if j == 6:
                    deposit2_arr.append("")
                    withdraw2_arr.append("")
                    expense_non_tax_arr.append("")
                if j == 4:
                    deposit2_arr.append("")
                    withdraw2_arr.append(data[i][9])
                    expense_non_tax_arr.append(data[i][5])
                if row[j] == "Exchange (optional)":
                    deposit2_arr.append(data[i][3]) 
                    withdraw2_arr.append("Uniswap V2 "+data[i][5])
                    expense_non_tax_arr.append("Uniswap V2 "+data[i][5])
                if row[j] == "Date":
                    deposit2_arr.append(data[i][0])
                    withdraw2_arr.append(data[i][0])
                    expense_non_tax_arr.append(data[i][0])
            general.append(withdraw2_arr)
            general.append(deposit2_arr)
            general.append(expense_non_tax_arr)
        
    return general



def main():
    row, types, value_dict = read_muster()
    general_save = []
    while True:
        #exchange = input("Von welcher Börse stammt die Datei? (Bitcoin.de = b)")
        exchange = "b"
        if exchange == "b":
            general = format_data_bitcoinde(row)
            general_save+=general
            for i in range(len(general_save)):
                print(f"{i+1}. {general_save[i]}")
            print('###############################################################################################################################################')
            print(f"Es wurden insgesamt {len(general_save)} Transaktionen gefunden.")
            print('###############################################################################################################################################')
        flag = input("Möchtest du noch mehr Dateien hochladen?(y/n)\n")
        if flag == "n":
            safety = input("Bitte schließ dein .xsl File in welches Sie das abspeichern wollen!! Drücke eine Taste, wenn du fortfahren willst.\n")
            save_to_xsl(general_save,row)
            break
        

if __name__ == '__main__':
    main()
