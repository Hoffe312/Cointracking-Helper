import pandas as pd
import openpyxl as op


def read_muster():
    df = pd.read_excel('C:\\Users\\LHP Borschel\\Desktop\\LHP HOFFMANN\\Persönliches\\CoinTracking_Excel_Import.xls')
    row_arr = df.iloc[0].values
    row_arr.tolist()
    type_arr = df.iloc[:, 0].tolist()
    type_arr = type_arr[1:]

    value_dict = {}
    for i in range(len(type_arr)):
        value_dict[type_arr[i]]= df.iloc[i+1,:].tolist()
    return row_arr, type_arr, value_dict

'''Funktion, welche für Abfragen, der unterschiedlichen Börsen verantwortlich ist'''
def path_format_handler():
    path = input("Gebe den Pfad zu deiner Datei an, welche du in das Cointracking_Excel_Import Format bringen willst:\n")
    #path= "C:\\Users\\LHP Borschel\\Desktop\\btc_account_statement_20130101-20131231.csv"
    data_format=""
    pattern_csv = ".csv"
    pattern_xsl = ".xsl"
    if pattern_csv in path:
        data_format = "c"

    if pattern_xsl in path:
        data_format = "x"
    return path, data_format


'''Liest excel oder csv Dateien ein und gibt dataframes zurück'''
def read_data(path,format):
    if path[0]=='"' and path[-1] =='"':
        path = path[1:-1]
    if format == "x":
        df = pd.read_excel(path)
        
    elif format == "c":
        df = pd.read_csv(path)
    df_array = []
    for i in range(len(df.iloc[:, 0].tolist())):
        df_row =df.iloc[i].values
        df_row =df_row[0].replace(";",", ")
        df_row_arr = df_row.split(",")
        df_array.append(df_row_arr)
    return df_array


def save_to_xsl(general,row):
    wb = op.Workbook()
    ws = wb.active
  
    for i, wert in enumerate(row):
        ws.cell(row=1, column=i+1).value = wert
    for i, row in enumerate(general):
        for j, value in enumerate(row):
            ws.cell(row=i+2, column=j+1).value = value
         
    wb.save("C:\\Users\\LHP Borschel\\Desktop\\LHP HOFFMANN\\Example.xlsx")
    print("******************************************")
    print("*DIE ERSTELLUNG DER DATEI WAR ERFOLGREICH*")
    print("******************************************")


'''Formatiert bitcoin.de dateien'''
def format_data_bitcoinde(row):
    path, format = path_format_handler()
    data = read_data(path,format)


    length_list = len(data)
    general = []
    #if general_array:
    #    general = general_array
    for i in range(length_list):

        if data[i][1] == " Verkauf":
            sell_arr = []
            for j in range(len(row)):
                if j == 0:
                    sell_arr.append("Trade")
                if j == 1:
                    zahl_string = data[i][8]
                    zahl_float = float(zahl_string.replace(",","."))
                    zahl_abs = abs(zahl_float)
                    sell_arr.append(zahl_abs)
                if j == 2:
                   sell_arr.append(data[i][9])
                if j == 3:
                    zahl_string = data[i][7]
                    zahl_float = float(zahl_string.replace(",","."))
                    zahl_abs = abs(zahl_float)
                    
                    sell_arr.append(zahl_abs)
                if j == 4:
                    sell_arr.append(data[i][2])
                if j == 5:
                    zahl_string = data[i][8]
                    zahl_float = float(zahl_string.replace(",","."))
                    zahl_abs = abs(zahl_float)
                    zahl_string2 = data[i][11]
                    zahl_float2 = float(zahl_string2.replace(",","."))
                    zahl_abs2 = abs(zahl_float2)
                    zahl = zahl_abs-zahl_abs2
                    sell_arr.append(zahl)
                if j == 6:
                    sell_arr.append(data[i][12])
                if j == 7:
                    sell_arr.append("bitcoin.de")
                if j == 8 or j == 9:
                    sell_arr.append("")
                if j == 10:
                    sell_arr.append(data[i][0])
            general.append(sell_arr)
        if data[i][1] == " Kauf":
            trade_arr = []
            for j in range(len(row)):
                
                if j == 0:
                    trade_arr.append("Trade")
                if j == 1:
                    zahl_string = data[i][7]
                    zahl_float = float(zahl_string.replace(",","."))
                    zahl_abs = abs(zahl_float)
                    
                    trade_arr.append(zahl_abs)
                if j == 2:
                    trade_arr.append(data[i][2])
                if j == 3:
                    zahl_string = data[i][8]
                    zahl_float = float(zahl_string.replace(",","."))
                    zahl_abs = abs(zahl_float)
                    trade_arr.append(zahl_abs)
                if j == 4:
                   trade_arr.append(data[i][9])
                if j == 5:
                    zahl_string = data[i][8]
                    zahl_float = float(zahl_string.replace(",","."))
                    zahl_abs = abs(zahl_float)
                    zahl_string2 = data[i][11]
                    zahl_float2 = float(zahl_string2.replace(",","."))
                    zahl_abs2 = abs(zahl_float2)
                    zahl = zahl_abs-zahl_abs2
                    trade_arr.append(zahl)
                if j == 6:
                    trade_arr.append(data[i][12])
                if j == 7:
                    trade_arr.append("bitcoin.de")
                if j == 8 or j == 9:
                    trade_arr.append("")
                if j == 10:
                    trade_arr.append(data[i][0])
            general.append(trade_arr)
        if str(data[i][1]) == " Auszahlung":
            withdraw_arr = [] 
            for j in range(len(row)):
                if j == 0:
                    withdraw_arr.append("Withdrawal")
                if j == 1 or j == 2 or  j == 8 or j == 9:
                    withdraw_arr.append("")
                if j == 5 and i < length_list-1:
                    for k in range(1,length_list-i):
                        if data[i+1*k][1] == " Netzwerk-Gebühr" and data[i+1*k][3] == data[i][3]:
                            zahl_string = data[i+1][13]
                            zahl_float = float(zahl_string.replace(",","."))
                            zahl_abs = abs(zahl_float)
                            withdraw_arr.append(zahl_abs)
                        else:
                            withdraw_arr.append("")
                        break
                if j == 6 and i < length_list-1:
                    if data[i+1][1] == " Netzwerk-Gebühr":
                        withdraw_arr.append(data[i+1][2])
                    else:
                        withdraw_arr.append("")

                if row[j] == "Sell Amount":
                    zahl_string = data[i][13]
                    zahl_float = float(zahl_string.replace(",","."))
                    zahl_abs = abs(zahl_float)
                    withdraw_arr.append(zahl_abs)
                if row[j] == "Sell Cur.":
                    withdraw_arr.append(data[i][2])
                if row[j] == "Exchange (optional)":
                    withdraw_arr.append("bitcoin.de")
                if row[j] == "Date":
                    
                    withdraw_arr.append(data[i][0])
            general.append(withdraw_arr)
    return general



def main():
    row, types, value_dict = read_muster()
    general_save = []
    while True:
        #exchange = input("Von welcher Börse stammt die Datei? (Bitcoin.de = b)")
        exchange = "b"
        if exchange == "b":
            general = format_data_bitcoinde(row)
            general_save+=general
            for i in range(len(general_save)):
                print(f"{i+1}. {general_save[i]}")
            print('###############################################################################################################################################')
            print(f"Es wurden insgesamt {len(general_save)} Transaktionen gefunden.")
            print('###############################################################################################################################################')
        flag = input("Möchtest du noch mehr Dateien hochladen?(y/n)\n")
        if flag == "n":
            safety = input("Bitte schließ dein .xsl File in welches Sie das abspeichern wollen!! Drücke eine Taste, wenn du fortfahren willst.\n")
            save_to_xsl(general_save,row)
            break
        

if __name__ == '__main__':
    main()
